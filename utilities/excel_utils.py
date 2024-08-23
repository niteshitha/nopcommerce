import openpyxl

def get_row_count(file,sheetname):      #Here in this function, 2 parameters are "filepath" and "sheetname"
    workBook = openpyxl.load_workbook(file)
    sheet = workBook[sheetname]
    return (sheet.max_row)        #This will return number of rows which contain test data(not empty cells)

def get_column_count(file,sheetname):      #Here 2 parameters are filepath and sheetname
    workBook = openpyxl.load_workbook(file)
    sheet = workBook[sheetname]
    return (sheet.max_column)        #This will return number of columns which contain test data

def read_data(file,sheetname,row_number,column_number):  #This function will read the specific data from specific cell
    workBook = openpyxl.load_workbook(file)
    sheet = workBook[sheetname]
    return(sheet.cell(row_number,column_number).value)

def write_data(file,sheetname,row_number,column_number,data):
    workBook = openpyxl.load_workbook(file)
    sheet = workBook[sheetname]
    sheet.cell(row_number,column_number).value = data
    workBook.save(file)



